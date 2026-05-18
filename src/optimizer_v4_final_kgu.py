"""
OPTIMIZER V4 FINAL - Using Real KGU Financial Model Data

Архітектура:
- Завантажує дані з KGU_Financial_Model.xlsx
- PV профіль: PVsyst аркуш (Col D = генерація по годинах)
- РДН ціни: Ретроспективні ціни аркуш (Col F = почасова ціна)
- Попит: Генерується синтетично (потребує реальних даних)
- Батарея параметри: Input аркуш
- Revenue: Врахує КПД 0.88 + РДН ціни + тарифи Укренерго

LP Формулювання (V4 = Strict PV Conservation):
- Variables: 192 (24h × 8)
- Constraints: 72 inequality + 48 equality
"""

import sqlite3
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List
import logging
import numpy as np
from scipy.optimize import linprog
import openpyxl

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

EXCEL_PATH = Path(__file__).parent.parent / "KGU_Financial_Model.xlsx"


class KGUDataLoader:
    """Завантажує дані з KGU Excel файлу"""
    
    def __init__(self, excel_path: Path):
        self.wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    def load_pv_profile(self, date_str: str) -> List[float]:
        """Завантажує PV профіль з PVsyst аркуша"""
        pv_sheet = self.wb['PVsyst']
        
        target_date = datetime.strptime(date_str, "%Y-%m-%d")
        ref_date = datetime(1990, 1, 1)
        
        day_offset = (target_date - ref_date).days
        day_in_year = (day_offset % 366)
        start_row = 19 + day_in_year * 24
        
        pv_values = []
        for hour in range(24):
            row = start_row + hour
            if row <= pv_sheet.max_row:
                pv_kw = pv_sheet.cell(row=row, column=4).value
                pv_values.append(float(pv_kw) if pv_kw else 0.0)
            else:
                pv_values.append(0.0)
        
        return pv_values
    
    def load_rdn_prices(self, date_str: str) -> List[float]:
        """Завантажує РДН ціни"""
        price_sheet = self.wb['Ретроспективні ціни']
        
        target_date = datetime.strptime(date_str, "%Y-%m-%d")
        ref_date = datetime(2024, 1, 1)
        
        day_offset = (target_date - ref_date).days
        day_in_file = (day_offset % 366)
        start_row = 3 + day_in_file * 24
        
        prices = []
        for hour in range(24):
            row = start_row + hour
            if row <= price_sheet.max_row:
                price = price_sheet.cell(row=row, column=6).value
                prices.append(float(price) if price else 3000.0)
            else:
                prices.append(3000.0)
        
        return prices
    
    def load_demand_profile(self, date_str: str) -> List[float]:
        """Завантажує реальний профіль попиту з demand_profile_2026.json"""
        import json
        
        demand_file = Path(__file__).parent.parent / "demand_profile_2026.json"
        
        if demand_file.exists():
            with open(demand_file, 'r') as f:
                data = json.load(f)
                hourly_demand = data.get('hourly_demand_kwh', [])
                
                if hourly_demand and len(hourly_demand) >= 8760:
                    target_date = datetime.strptime(date_str, "%Y-%m-%d")
                    ref_date = datetime(2026, 1, 1)
                    day_offset = (target_date - ref_date).days
                    day_in_year = (day_offset % 365)
                    start_hour = day_in_year * 24
                    end_hour = start_hour + 24
                    return hourly_demand[start_hour:end_hour]
        
        logger.warning(f"Не знайдено реальний попит, використовую синтетичний")
        daily_demand = 170.0
        hourly_factor = [
            0.5 if h < 6 else
            1.2 if 6 <= h < 18 else
            0.8
            for h in range(24)
        ]
        sum_factor = sum(hourly_factor)
        return [f * daily_demand / sum_factor for f in hourly_factor]


class Optimizer24hV4Final:
    """V4 LP Optimizer з реальними KGU даними"""
    
    def __init__(self, excel_path: Path = EXCEL_PATH):
        self.loader = KGUDataLoader(excel_path)
        self.config = {
            "pv_capacity_kw": 8699.6,
            "battery_capacity_kwh": 10032,
            "battery_soc_min_percent": 20,
            "battery_soc_max_percent": 80,
            "battery_efficiency_round_trip": 0.88,
            "battery_max_charge_kw": 2500,
            "battery_max_discharge_kw": 2200,
            "grid_max_import_kw": 5000,
            "grid_max_export_kw": 5000,
        }
    
    def optimize_day(self, date: str, initial_soc_kwh: float = 5016.0) -> Dict:
        """Оптимізує один день"""
        
        prices = self.loader.load_rdn_prices(date)
        pv_profile = self.loader.load_pv_profile(date)
        demand_profile = self.loader.load_demand_profile(date)
        
        logger.info(f"[V4-FINAL] Оптимізую {date}")
        logger.info(f"  PV: max={max(pv_profile):.1f} кВт")
        logger.info(f"  Price: min={min(prices):.0f}, max={max(prices):.0f} грн/MWh")
        
        n_hours = 24
        n_vars_per_hour = 8
        n_vars = n_hours * n_vars_per_hour
        
        def idx(h, var_type):
            return h * n_vars_per_hour + var_type
        
        c = np.zeros(n_vars)
        tariff_dist = 686.23
        tariff_disp = 98.97
        
        for h in range(n_hours):
            c[idx(h, 6)] -= prices[h] / 1000
            c[idx(h, 5)] += (prices[h] + tariff_dist + tariff_disp) / 1000
            c[idx(h, 4)] += (1 - self.config["battery_efficiency_round_trip"]) * prices[h] / 1000
        
        bounds = []
        for h in range(n_hours):
            bounds.append((0, pv_profile[h]))
            bounds.append((0, pv_profile[h]))
            bounds.append((0, pv_profile[h]))
            bounds.append((0, self.config["battery_max_charge_kw"]))
            bounds.append((0, self.config["battery_max_discharge_kw"]))
            bounds.append((0, self.config["grid_max_import_kw"]))
            bounds.append((0, self.config["grid_max_export_kw"]))
            bounds.append((
                self.config["battery_capacity_kwh"] * self.config["battery_soc_min_percent"] / 100,
                self.config["battery_capacity_kwh"] * self.config["battery_soc_max_percent"] / 100
            ))
        
        A_ub_list = []
        b_ub_list = []
        A_eq_list = []
        b_eq_list = []
        
        for h in range(n_hours):
            # Energy balance
            row = np.zeros(n_vars)
            row[idx(h, 0)] = -1
            row[idx(h, 4)] = -1
            row[idx(h, 5)] = -1
            A_ub_list.append(row)
            b_ub_list.append(-demand_profile[h])
            
            # PV conservation
            row = np.zeros(n_vars)
            row[idx(h, 0)] = 1
            row[idx(h, 1)] = 1
            row[idx(h, 2)] = 1
            A_eq_list.append(row)
            b_eq_list.append(pv_profile[h])
            
            # Grid export constraint
            row = np.zeros(n_vars)
            row[idx(h, 6)] = 1
            row[idx(h, 1)] = -1
            row[idx(h, 4)] = -self.config["battery_efficiency_round_trip"]
            A_ub_list.append(row)
            b_ub_list.append(0)
            
            # Battery charge constraint
            row = np.zeros(n_vars)
            row[idx(h, 3)] = 1
            row[idx(h, 2)] = -1
            A_ub_list.append(row)
            b_ub_list.append(0)
            
            # SoC dynamics
            row = np.zeros(n_vars)
            row[idx(h, 3)] = -self.config["battery_efficiency_round_trip"]
            row[idx(h, 4)] = 1.0 / self.config["battery_efficiency_round_trip"]
            row[idx(h, 7)] = 1
            
            if h == 0:
                A_eq_list.append(row)
                b_eq_list.append(initial_soc_kwh)
            else:
                row[idx(h-1, 7)] = -1
                A_eq_list.append(row)
                b_eq_list.append(0)
        
        A_ub = np.array(A_ub_list) if A_ub_list else None
        b_ub = np.array(b_ub_list) if b_ub_list else None
        A_eq = np.array(A_eq_list) if A_eq_list else None
        b_eq = np.array(b_eq_list) if b_eq_list else None
        
        try:
            result = linprog(
                c, A_ub=A_ub, b_ub=b_ub, A_eq=A_eq, b_eq=b_eq,
                bounds=bounds, method='highs', options={'disp': False}
            )
            
            if not result.success:
                logger.warning(f"LP failed: {result.message}")
                return {}
            
            x = result.x
        except Exception as e:
            logger.error(f"LP error: {e}")
            return {}
        
        dispatch = []
        total_revenue = 0.0
        
        for h in range(n_hours):
            revenue = (
                x[idx(h, 6)] * prices[h] / 1000 -
                x[idx(h, 5)] * (prices[h] + tariff_dist + tariff_disp) / 1000 -
                x[idx(h, 4)] * (1 - self.config["battery_efficiency_round_trip"]) * prices[h] / 1000
            )
            total_revenue += revenue
            
            dispatch.append({
                "hour": h,
                "pv_to_demand": x[idx(h, 0)],
                "pv_to_export": x[idx(h, 1)],
                "pv_to_charge": x[idx(h, 2)],
                "batt_charge": x[idx(h, 3)],
                "batt_discharge": x[idx(h, 4)],
                "grid_import": x[idx(h, 5)],
                "grid_export": x[idx(h, 6)],
                "soc": x[idx(h, 7)],
                "revenue": revenue,
            })
        
        logger.info(f"  Revenue: {total_revenue:,.0f} грн")
        
        return {
            "date": date,
            "total_revenue": total_revenue,
            "final_soc": x[idx(23, 7)],
            "dispatch": dispatch,
        }


if __name__ == "__main__":
    opt = Optimizer24hV4Final()
    result = opt.optimize_day("2024-01-15")
    if result:
        print(f"✅ {result['date']} → Revenue: {result['total_revenue']:,.0f} грн")
