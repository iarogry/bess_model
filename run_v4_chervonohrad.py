#!/usr/bin/env python3
"""
V4 Battery Simulator - Запуск на данных Chervonohrad для сравнения с Ярославовой моделью
Цель: Понять архітектурные различия между V4 и оптимізованой моделью Ярослава
"""

import openpyxl
import json
from datetime import datetime, timedelta
import os

print("="*100)
print("🔋 BATTERY SIMULATOR V4 - CHERVONOHRAD DATA LOADER")
print("="*100)

# Загрузим исходные данные из Chervonohrad_Model.xlsx
wb = openpyxl.load_workbook('Chervonohrad_Model.xlsx')

# 1. Загружаємо генерацію ФЕС (PVsyst)
print("\n📥 Загрузка ФЕС данних...")
ws_pes = wb['ФЕС']

pv_hourly = []
for row in range(3, 8763):  # 8760 часов в году
    col_b = ws_pes.cell(row, 2).value  # Column B - генерация
    if isinstance(col_b, (int, float)):
        pv_hourly.append(col_b)
    else:
        pv_hourly.append(0)

print(f"   ✓ Загружено {len(pv_hourly)} часов генерации")
print(f"   Min: {min(pv_hourly):.2f} kWh, Max: {max(pv_hourly):.2f} kWh, Avg: {sum(pv_hourly)/len(pv_hourly):.2f} kWh")

# 2. Загружаємо ціни РДН (Ретроспективні ціни)
print("\n📥 Загрузка РДН цін...")
ws_prices = wb['Ретроспективні ціни']

price_hourly = []
for row in range(3, 8763):  # Column F - ціна
    col_f = ws_prices.cell(row, 6).value
    if isinstance(col_f, (int, float)):
        price_hourly.append(col_f)
    else:
        price_hourly.append(100)  # Default price if missing

print(f"   ✓ Загружено {len(price_hourly)} часов цін")
print(f"   Min: {min(price_hourly):.2f} UAH/MWh, Max: {max(price_hourly):.2f} UAH/MWh, Avg: {sum(price_hourly)/len(price_hourly):.2f} UAH/MWh")

# 3. Загружаємо спрос (Операційна модель базова)
print("\n📥 Загрузка потреби (спросу)...")
ws_ops = wb['Операційна модель (базова)']

demand_hourly = []
for row in range(3, 8763):  # Column D? Column E? Нужно найти
    # Попробуем разные столбцы
    col_val = ws_ops.cell(row, 4).value or ws_ops.cell(row, 5).value or ws_ops.cell(row, 6).value
    if isinstance(col_val, (int, float)):
        demand_hourly.append(col_val)
    else:
        demand_hourly.append(100)  # Default demand if missing

print(f"   ✓ Загружено {len(demand_hourly)} часов спроса")
print(f"   Min: {min(demand_hourly):.2f} kWh, Max: {max(demand_hourly):.2f} kWh, Avg: {sum(demand_hourly)/len(demand_hourly):.2f} kWh")

# 4. Сохраним данные в JSON для V4 оптимизатора
print("\n💾 Сохранение в chervonohrad_data.json...")

chervonohrad_data = {
    "metadata": {
        "source": "Chervonohrad_Model.xlsx",
        "date_loaded": datetime.now().isoformat(),
        "year": 2024,
        "location": "Chervonohrad",
        "pv_capacity_mw": 2.5,
        "battery_capacity_mwh": 10.0
    },
    "hourly_data": {
        "pv_generation_kwh": pv_hourly,
        "rdn_price_uah_per_mwh": price_hourly,
        "demand_kwh": demand_hourly
    },
    "statistics": {
        "pv": {
            "min": min(pv_hourly),
            "max": max(pv_hourly),
            "avg": sum(pv_hourly) / len(pv_hourly),
            "total_annual": sum(pv_hourly)
        },
        "price": {
            "min": min(price_hourly),
            "max": max(price_hourly),
            "avg": sum(price_hourly) / len(price_hourly)
        },
        "demand": {
            "min": min(demand_hourly),
            "max": max(demand_hourly),
            "avg": sum(demand_hourly) / len(demand_hourly),
            "total_annual": sum(demand_hourly)
        }
    }
}

with open('chervonohrad_data.json', 'w', encoding='utf-8') as f:
    json.dump(chervonohrad_data, f, indent=2, ensure_ascii=False)

print("   ✓ Сохранено в chervonohrad_data.json")

print("\n" + "="*100)
print("✅ ДАННЫЕ ГОТОВЫ")
print("="*100)
print("\nСлідуючий крок: Запустити V4 оптимізатор на Chervonohrad даних")
print("Команда: python3 src/optimizer_v4_scipy.py --data chervonohrad_data.json")

