#!/usr/bin/env python3
"""
V4 Battery Simulator - Chervonohrad Data (CORRECTED)
Загрузка реальных данных из Chervonohrad_Model.xlsx для сравнения
"""

import openpyxl
import json
from datetime import datetime
import sys

print("="*100)
print("🔋 BATTERY SIMULATOR V4 - CHERVONOHRAD DATA LOADER (CORRECTED)")
print("="*100)

wb = openpyxl.load_workbook('Chervonohrad_Model.xlsx')

# ЗАГРУЖАЕМ ЦІНИ РДН (столбец I из Операційна модель)
print("\n📥 Загрузка РДН цін (Column I)...")
ws = wb['Операційна модель (базова)']

prices_rdn = []
for row in range(12, 8772):  # Строки 12-8771 = 8760 часов
    val = ws.cell(row, 9).value  # Column I (9) = Ціна РДН
    if isinstance(val, (int, float)):
        prices_rdn.append(val)
    else:
        # Попробуем получить значение формулы через промежуточный файл
        prices_rdn.append(3000)  # Default средняя цена

print(f"   ✓ Загружено {len(prices_rdn)} часов РДН цін")
if prices_rdn:
    print(f"     Min: {min(prices_rdn):.2f}, Max: {max(prices_rdn):.2f}, Avg: {sum(prices_rdn)/len(prices_rdn):.2f} UAH/MWh")

# ЗАГРУЖАЄМО ЦЕНЫ ИЗ ЛИСТА "Ретроспективні ціни" (более надежный источник)
print("\n📥 Загрузка РДН цін из 'Ретроспективні ціни' (Column F)...")
ws = wb['Ретроспективні ціни']

prices_historical = []
for row in range(3, 8763):  # Строки 3-8762 = 8760 часов
    val = ws.cell(row, 6).value  # Column F = РДН ціна
    if isinstance(val, (int, float)):
        prices_historical.append(val)
    else:
        prices_historical.append(0)

print(f"   ✓ Загружено {len(prices_historical)} часов")
if prices_historical:
    # Фильтруем нули
    valid_prices = [p for p in prices_historical if p > 0]
    print(f"     Valid prices: {len(valid_prices)}")
    if valid_prices:
        print(f"     Min: {min(valid_prices):.2f}, Max: {max(valid_prices):.2f}, Avg: {sum(valid_prices)/len(valid_prices):.2f} UAH/MWh")

# ГЕНЕРАЦИЯ ФЕС (Column H из Операційна модель)
print("\n📥 Загрузка генерації ФЕС (Column H)...")
ws = wb['Операційна модель (базова)']

pv_generation = []
for row in range(12, 8772):
    val = ws.cell(row, 8).value  # Column H = Генерація ФЕС
    if isinstance(val, (int, float)):
        pv_generation.append(val)
    else:
        pv_generation.append(0)

print(f"   ✓ Загружено {len(pv_generation)} часов ФЕС")
if pv_generation:
    valid_pv = [p for p in pv_generation if p > 0]
    print(f"     Valid generation hours: {len(valid_pv)}")
    if valid_pv:
        print(f"     Min: {min(valid_pv):.2f}, Max: {max(valid_pv):.2f}, Avg: {sum(valid_pv)/len(valid_pv):.2f} kWh")

# ИСПОЛЬЗУЕМЫЕ ОБЪЕМЫ ТОРГОВЛИ (Column S = Продаж, Column T = Купівля)
print("\n📥 Загрузка фактических объемов торговли на РДН...")
ws = wb['Операційна модель (базова)']

volumes_sold = []
volumes_bought = []
for row in range(12, 8772):
    sold = ws.cell(row, 19).value  # Column S = Продаж на РДН [кВт*год]
    bought = ws.cell(row, 20).value  # Column T = Купівля на РДН [кВт*год]
    
    if isinstance(sold, (int, float)):
        volumes_sold.append(sold)
    else:
        volumes_sold.append(0)
    
    if isinstance(bought, (int, float)):
        volumes_bought.append(bought)
    else:
        volumes_bought.append(0)

print(f"   ✓ Загружено {len(volumes_sold)} часов")
print(f"     Продажі: Sum={sum(volumes_sold):.2f} kWh, Avg={sum(volumes_sold)/len(volumes_sold):.4f} kWh")
print(f"     Покупки: Sum={sum(volumes_bought):.2f} kWh, Avg={sum(volumes_bought)/len(volumes_bought):.4f} kWh")

# СОХРАНЯЕМ В JSON
print("\n💾 Сохранение данных в chervonohrad_data_corrected.json...")

data = {
    "metadata": {
        "source": "Chervonohrad_Model.xlsx - Операційна модель (базова)",
        "date_loaded": datetime.now().isoformat(),
        "year": 2024,
        "location": "Chervonohrad",
        "project": "2.5 MW PV + 10 MWh BESS",
        "note": "РДН торговля (без локального спроса)"
    },
    "hourly_data": {
        "price_rdn_uah_per_mwh": prices_historical,
        "pv_generation_kwh": pv_generation,
        "volumes_sold_kwh": volumes_sold,
        "volumes_bought_kwh": volumes_bought
    },
    "statistics": {
        "prices": {
            "count": len([p for p in prices_historical if p > 0]),
            "min": min([p for p in prices_historical if p > 0]) if any(p > 0 for p in prices_historical) else 0,
            "max": max([p for p in prices_historical if p > 0]) if any(p > 0 for p in prices_historical) else 0,
            "avg": sum([p for p in prices_historical if p > 0]) / len([p for p in prices_historical if p > 0]) if any(p > 0 for p in prices_historical) else 0
        },
        "pv": {
            "total": sum(pv_generation),
            "min": min([p for p in pv_generation if p > 0]) if any(p > 0 for p in pv_generation) else 0,
            "max": max(pv_generation),
            "avg": sum(pv_generation) / len(pv_generation)
        }
    }
}

with open('chervonohrad_data_corrected.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, indent=2, ensure_ascii=False)

print("   ✓ Сохранено в chervonohrad_data_corrected.json")

print("\n" + "="*100)
print("✅ ДАННЫЕ ЗАГРУЖЕНЫ И ГОТОВЫ")
print("="*100)

print("\n📊 АРХИТЕКТУРНЫЙ ВЫВОД:")
print("   • Ярославова модель оптимизирует ТОЛЬКО торговлю на РДН")
print("   • Спроса подприємства в этой модели НЕ учитывается")
print("   • V4 должен быть адаптирован для чистой РДН торговли (без локального спроса)")
print("\n🚀 Следующий шаг: Создать V4-RDN оптимизатор (чистая РДН торговля без спроса)")

